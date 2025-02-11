import streamlit as st
import pandas as pd
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime
from typing import List

# =============================================================================
# App Configuration
# =============================================================================
st.set_page_config(
    layout="wide",
    page_title="Budget",
    page_icon="logo.png"  # Uses logo.png as the page icon.
)

st.markdown(
    """
    <style>
    .block-container {
        padding-top: 50px;
    }
    [data-testid="stSidebar"] > div:first-child {
        padding-top: 10px;
    }
    [data-testid="stSidebar"] div.stButton button {
        width: 100% !important;
    }
    </style>
    """,
    unsafe_allow_html=True
)


# =============================================================================
# ExcelApp Class Definition
# =============================================================================
class ExcelApp:
    def __init__(self, file_path: str, headers: List[str]):
        self.file_path = file_path
        self.headers = headers
        self.ensure_file()
        self.load_data()

    def ensure_file(self) -> None:
        """Ensure the Excel file exists with a 'Main' sheet that has the headers."""
        if not os.path.exists(self.file_path):
            wb = Workbook()
            ws = wb.active
            ws.title = "Main"
            ws.append(self.headers)
            wb.save(self.file_path)
        else:
            try:
                wb = load_workbook(self.file_path)
                if "Main" not in wb.sheetnames:
                    ws = wb.create_sheet("Main")
                    ws.append(self.headers)
                    wb.save(self.file_path)
            except Exception as e:
                st.error(f"Error ensuring file: {e}")

    def load_data(self) -> None:
        """Load data from the Excel file into st.session_state.df."""
        try:
            df = pd.read_excel(self.file_path, sheet_name="Main", engine="openpyxl")
        except Exception as e:
            st.error(f"Error loading data: {e}")
            df = pd.DataFrame(columns=self.headers)
        st.session_state.df = df

    def save_data(self) -> None:
        """Save st.session_state.df to the Excel file (overwriting it)."""
        try:
            with pd.ExcelWriter(self.file_path, engine="openpyxl", mode="w") as writer:
                st.session_state.df.to_excel(writer, sheet_name="Main", index=False)
            st.success("✅ Data saved successfully!")
        except Exception as e:
            st.error(f"Error saving data: {e}")

    def add_row(self) -> None:
        """Append a new blank row to the DataFrame."""
        new_row = pd.DataFrame([[None] * len(self.headers)], columns=self.headers)
        st.session_state.df = pd.concat([st.session_state.df, new_row], ignore_index=True)

    def clear_data(self) -> None:
        """Clear all data (set DataFrame to empty with headers)."""
        st.session_state.df = pd.DataFrame(columns=self.headers)

    def show_editor(self) -> None:
        """Display the data editor widget and update st.session_state.df."""
        st.subheader("📝 Data Editor")
        if st.session_state.df.empty:
            st.session_state.df = pd.DataFrame(columns=self.headers)
        # st.data_editor returns the updated DataFrame
        updated_df = st.data_editor(
            st.session_state.df,
            key="data_editor",
            num_rows="dynamic",
            use_container_width=True,
            hide_index=True,
            column_config={
                "Account Number": st.column_config.NumberColumn(
                    "Account Number", width="medium", step=1, min_value=0
                ),
                "Account Name": st.column_config.TextColumn(
                    "Account Name", width="medium"
                ),
                **{
                    month: st.column_config.NumberColumn(
                        month, width="small", format="%.2f", min_value=0, required=False
                    )
                    for month in self.headers[2:]
                }
            }
        )
        st.session_state.df = updated_df

    def show_excel(self) -> None:
        """Display a read-only view of the Excel file read from disk."""
        st.subheader("📊 Excel Sheet")
        try:
            df = pd.read_excel(self.file_path, sheet_name="Main", engine="openpyxl")
            st.dataframe(df, use_container_width=True)
        except Exception as e:
            st.error(f"Error reading Excel file: {e}")


# =============================================================================
# Initialization
# =============================================================================
headers = [
    "Account Number", "Account Name",
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December"
]
file_path = "example.xlsx"
app = ExcelApp(file_path, headers)

# Initialize view toggle if not already set.
if "view" not in st.session_state:
    st.session_state.view = "editor"  # or "excel"
if "show_instructions" not in st.session_state:
    st.session_state.show_instructions = False

# =============================================================================
# Sidebar
# =============================================================================
with st.sidebar:
    st.image("logo.png", use_container_width=True)
    st.markdown("<br>", unsafe_allow_html=True)

    # Toggle Instructions
    instr_label = "Hide Instructions" if st.session_state.show_instructions else "Show Instructions"
    if st.button(instr_label):
        st.session_state.show_instructions = not st.session_state.show_instructions
    if st.session_state.show_instructions:
        st.info(
            """
            **Instructions:**
            1. Edit cells by clicking on them and then press Enter or click outside to commit changes.
            2. Use the Add Row button to insert a new blank row.
            3. Click Save to write changes to the Excel file.
            4. Click Clear Data to remove all rows.
            5. Toggle between the Editor and Excel view using the Toggle View button.
            """
        )

    # Toggle between Editor and Excel view.
    view_label = "Show Excel Sheet" if st.session_state.view == "editor" else "Show Editor"
    if st.button(view_label):
        st.session_state.view = "excel" if st.session_state.view == "editor" else "editor"

    # Control Buttons
    if st.button("➕ Add Row"):
        app.add_row()
    if st.button("💾 Save"):
        app.save_data()
    if st.button("🗑️ Clear Data"):
        app.clear_data()

# =============================================================================
# Main Area
# =============================================================================
if st.session_state.view == "editor":
    app.show_editor()
else:
    app.show_excel()
