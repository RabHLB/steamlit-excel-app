import os
import platform
from openpyxl import load_workbook
from datetime import datetime

# Load the existing Excel file
file_path = "example.xlsx"
wb = load_workbook(file_path)
ws = wb.active

# Find the next empty row
next_row = ws.max_row + 1  # Gets the next available row

# Get user input
col_a_input = input("Enter value for column A: ")
col_b_input = input("Enter value for column B: ")

# Get current timestamp
timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")  # Format: YYYY-MM-DD HH:MM:SS

# Write user input and timestamp to the next available row
ws[f"A{next_row}"] = col_a_input
ws[f"B{next_row}"] = col_b_input
ws[f"C{next_row}"] = timestamp  # Write timestamp automatically

# Save and close the file
wb.save(file_path)
wb.close()

print(f"Data added successfully to row {next_row} with timestamp {timestamp} in {file_path}!")

# Automatically open the file after saving
if platform.system() == "Windows":
    os.startfile(file_path)  # Windows
elif platform.system() == "Darwin":  # macOS
    os.system(f"open {file_path}")
else:  # Linux
    os.system(f"xdg-open {file_path}")
